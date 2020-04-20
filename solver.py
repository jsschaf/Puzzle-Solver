import pandas as pd
import numpy as np
import itertools
import openpyxl

# SUDOKU: find next cell to populate
def find_empty(puzzle):
    for i in range(9):
        for j in range(9):
            if puzzle[i][j] == 0:
                return (i, j)  # row, col

    return None

# SUDOKU: check if current number can exist in cell at ind
def is_valid(puzzle, num, ind):
    i, j = ind

    # check col
    for col in range(9):
        if col is not j:
            if num == puzzle[i][col]:
                return False

    # check row
    for row in range(9):
        if row is not i:
            if num == puzzle[row][j]:
                return False

    # check box
    box_i = j // 3
    box_j = i // 3

    for row in range(box_j*3, box_j*3+3):
        for col in range(box_i*3, box_i*3+3):
            if (row, col) != ind:
                if puzzle[row, col] == num:
                    return False

    return True

# SUDOKU: solve puzzle recursively
def solve(puzzle):
    ind = find_empty(puzzle)
    if not ind:
        return True # no more empty slots
    i, j = ind

    for num in range(1, 10):
        if is_valid(puzzle, num, ind):
            puzzle[i][j] = num
            if solve(puzzle):
                return True
            
            puzzle[i][j] = 0
    return False

# Read in data for both puzzles
df = pd.read_excel('puzzle_template.xlsx', skip_rows=1, nrows=9, usecols="B:J")
crunch_df = pd.read_excel('puzzle_template.xlsx', skip_rows=1, nrows=6, usecols="O:T")
crunch = crunch_df.to_numpy()
puzzle = df.to_numpy()

# CRUNCH: create rows and columns data structures
crunch_rows = crunch[::2]
crunch_trans = crunch.transpose()
crunch_cols = crunch_trans[::2]
perms = list(itertools.permutations(range(1, 10)))
solution = []
#try each combination
for elt in perms:

    # check if first row works
    answer = eval(str(elt[0]) + crunch_rows[0][1] + str(elt[1]) + crunch_rows[0][3] + str(elt[2]))
    if answer == crunch_rows[0][5]:
        # solve second row
        answer = eval(str(elt[3]) + crunch_rows[1][1] + str(elt[4]) + crunch_rows[1][3] + str(elt[5]))
        if answer == crunch_rows[1][5]:
            # solve last row
            answer = eval(str(elt[6]) + crunch_rows[2][1] + str(elt[7]) + crunch_rows[2][3] + str(elt[8]))
            if answer == crunch_rows[2][5]:
                #move on to columns
                answer = eval(str(elt[0]) + crunch_cols[0][1] + str(elt[3]) + crunch_cols[0][3] + str(elt[6]))
                if answer == crunch_cols[0][5]:
                    answer = eval(str(elt[1]) + crunch_cols[1][1] + str(elt[4]) + crunch_cols[1][3] + str(elt[7]))
                    if answer == crunch_cols[1][5]:
                        answer = eval(str(elt[2]) + crunch_cols[2][1] + str(elt[5]) + crunch_cols[2][3] + str(elt[8]))
                        if answer == crunch_cols[2][5]:
                            solution = elt
                            break

# SUDOKU: make sure we have a 9x9 puzzle
rows, cols = puzzle.shape
newcol = np.array([0, 0, 0, 0, 0, 0, 0, 0, 0]) 
while cols is not 9:
    puzzle = np.column_stack((puzzle, newcol))
    cols += 1
while rows is not 9:
    puzzle = np.vstack((puzzle, newcol))
    rows += 1

for i in range(9):
    for j in range(9):
        if puzzle[i][j] != puzzle[i][j]:
            puzzle[i][j] = 0

solve(puzzle)
solved = pd.DataFrame(puzzle)

wb = openpyxl.load_workbook("puzzle_template.xlsx")
ws = wb.get_sheet_by_name('Sheet1')

for i in range(2, 11):
    for j in range(2, 11):

        ws.cell(row=i, column=j).value = puzzle[i-2][j-2]

#fill in number crunch cells:

ws.cell(row=2, column=15).value = solution[0]
ws.cell(row=2, column=17).value = solution[1]
ws.cell(row=2, column=19).value = solution[2]
ws.cell(row=4, column=15).value = solution[3]
ws.cell(row=4, column=17).value = solution[4]
ws.cell(row=4, column=19).value = solution[5]
ws.cell(row=6, column=15).value = solution[6]
ws.cell(row=6, column=17).value = solution[7]
ws.cell(row=6, column=19).value = solution[8]

wb.save("solved_puzzles.xlsx")